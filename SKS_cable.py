import streamlit as st
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Border, Side, Alignment

def process_data(FileName, FileName_Sheet):

    # Подготовка таблицы к работе
    df = pd.read_excel(FileName, sheet_name=FileName_Sheet)
    df = df.iloc[0:, 1:]
    df.columns = ['№  кабеля', 'Марка кабеля', 'Жильность x сечение', 'Кол-во использ. Жил', 'Откуда', 'Куда',
                  'Длина проект, м', 'Длина факт, м', 'Примечание']

    df.loc[df['Длина проект, м'].isna(), ['Длина проект, м']] = 0
    df.loc[df['Длина факт, м'].isna(), ['Длина факт, м']] = 0
    df = df.dropna(ignore_index=True)

    # Обработка столбцов Откуда и Куда, сокращение до названия шкафа
    # формируем таблицу для работы
    df_2 = df.copy()
    df_2 = df_2.drop(['Марка кабеля', 'Марка кабеля', 'Жильность x сечение',
                      'Кол-во использ. Жил', 'Длина проект, м', 'Длина факт, м', 'Примечание'], axis=1)
    df_2['№  кабеля'] = df_2['№  кабеля'].astype(str)
    df_2['Откуда'] = df_2['Откуда'].astype(str)
    df_2['Куда'] = df_2['Куда'].astype(str)

    # Убираем непечатные символы
    df_2 = df_2.apply(lambda x: x.str.replace(r'[\r\n\t]', '', regex=True) if x.dtype == "str" else x)

    # убираем скобки перед и после номера кабеля
    df_2['№  кабеля'] = df_2['№  кабеля'].str.replace('(','').replace(')','')

    # выделяем номер шкафа
    df_2['Откуда'] = '=' + df_2['Откуда'].str.partition('=')[2]
    df_2['Откуда'] = df_2['Откуда'].str.partition('-')[0]
    df_2['Куда'] = '=' + df_2['Куда'].str.partition('=')[2]
    df_2['Куда'] = df_2['Куда'].str.partition('-')[0]

    # удаление лишних пробелов в названии шкафа
    df_2['Откуда'] = df_2['Откуда'].str.replace(' ', '')
    df_2['Куда'] = df_2['Куда'].str.replace(' ', '')

    # Удаление лишних пробелов в конце названия шкафа
    df['Откуда'] = df['Откуда'].str.rstrip()
    df['Куда'] = df['Куда'].str.rstrip()

    # Формирование таблицы для проверки соединений кабеля
    # Удаляем из таблицы кабель внутри шкафов

    df_cab = df_2.loc[df_2['Откуда'] != df_2['Куда']]

    # Переносит шкафы W1, W2, W3, W4 из Куда в Откуда для удобства чтения и сопртировки
    # формируем итоговую таблицу

    df_cabr = df_cab.loc[(df_2['Откуда'] == '=X00+W1')]
    df_cabx = df_cab.loc[(df_2['Куда'] == '=X00+W1')]
    df_cabx = df_cabx[['№  кабеля', 'Куда', 'Откуда']]
    df_cabx.columns = ['№  кабеля', 'Откуда', 'Куда']
    df_cabr = pd.concat([df_cabr, df_cabx])

    df_cab = df_cab.loc[(df_2['Откуда'] != '=X00+W1')]
    df_cab = df_cab.loc[(df_2['Куда'] != '=X00+W1')]

    df_cabr = pd.concat([df_cabr, df_cab.loc[(df_2['Откуда'] == '=X00+W2')]])
    df_cabx = df_cab.loc[(df_2['Куда'] == '=X00+W2')]
    df_cabx = df_cabx[['№  кабеля', 'Куда', 'Откуда']]
    df_cabx.columns = ['№  кабеля', 'Откуда', 'Куда']
    df_cabr = pd.concat([df_cabr, df_cabx])

    df_cab = df_cab.loc[(df_2['Откуда'] != '=X00+W2')]
    df_cab = df_cab.loc[(df_2['Куда'] != '=X00+W2')]

    df_cabr = pd.concat([df_cabr, df_cab.loc[(df_2['Откуда'] == '=X00+W3')]])
    df_cabx = df_cab.loc[(df_2['Куда'] == '=X00+W3')]
    df_cabx = df_cabx[['№  кабеля', 'Куда', 'Откуда']]
    df_cabx.columns = ['№  кабеля', 'Откуда', 'Куда']
    df_cabr = pd.concat([df_cabr, df_cabx])

    df_cab = df_cab.loc[(df_2['Откуда'] != '=X00+W3')]
    df_cab = df_cab.loc[(df_2['Куда'] != '=X00+W3')]

    df_cabr = pd.concat([df_cabr, df_cab.loc[(df_2['Откуда'] == '=X00+W4')]])
    df_cabx = df_cab.loc[(df_2['Куда'] == '=X00+W4')]
    df_cabx = df_cabx[['№  кабеля', 'Куда', 'Откуда']]
    df_cabx.columns = ['№  кабеля', 'Откуда', 'Куда']
    df_cabr = pd.concat([df_cabr, df_cabx])

    df_cab = df_cab.loc[(df_2['Откуда'] != '=X00+W4')]
    df_cab = df_cab.loc[(df_2['Куда'] != '=X00+W4')]

    # обработаем оставшиеся кабели
    df_cab2 = df_cab[['№  кабеля', 'Куда', 'Откуда']]
    df_cab2.columns = ['№  кабеля', 'Откуда', 'Куда']

    df_cab = pd.concat([df_cab, df_cab2])
    df_cab = df_cab.sort_values(by='Откуда')
    df_cab = df_cab.drop_duplicates(subset=['№  кабеля'])

    # соединяем таблицы
    df_cabr = pd.concat([df_cabr, df_cab], ignore_index=True)

    # Сортируем таблицу по номеру кабеля
    df_cabr.sort_values(by='№  кабеля', inplace=True)

    # выгружаем результат в файл Excel
    wb = Workbook()
    ws = wb.active

    # добавим пустой пробел перед знаком =, чтобы Excel не считал значение формулой
    df_cabr['№  кабеля'] = '\u200b' + df_cabr['№  кабеля']
    df_cabr['Откуда'] = '\u200b' + df_cabr['Откуда']
    df_cabr['Куда'] = '\u200b' + df_cabr['Куда']

    # заполняем ячейки из таблицы
    for r in dataframe_to_rows(df_cabr, index=False, header=True):
        ws.append(r)

    # зададим ширину столбцов 20 единиц
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20

    # зададим стиль первого ряда - заголовка таблицы
    # Создание стиля шрифта - жирный
    bold_font = Font(bold=True, color="000000")  # Жирный, черный

    # Применение стилей к ячейке
    cell_A1 = ws['A1']
    cell_A1.font = bold_font
    cell_A1.alignment = Alignment(horizontal='center', vertical='center')  # Центрирование

    cell_B1 = ws['B1']
    cell_B1.font = bold_font
    cell_B1.alignment = Alignment(horizontal='center', vertical='center')  # Центрирование

    cell_C1 = ws['C1']
    cell_C1.font = bold_font
    cell_C1.alignment = Alignment(horizontal='center', vertical='center')  # Центрирование

    # запись результата в буфер
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

# --- Интерфейс Streamlit ---
st.title("Таблица СКС")

uploaded_file = st.file_uploader("Загрузите Кабельный журнал в формате Excel", type=['xlsx'])

if uploaded_file is not None:
    wb_temp = load_workbook(uploaded_file, read_only=True)
    sheets = wb_temp.sheetnames
    selected_sheet = st.selectbox("Выберите лист:", sheets)

    if st.button("✨ Обработать файл"):
        # Важно: для openpyxl нужно сбросить указатель файла
        uploaded_file.seek(0)
        result = process_data(uploaded_file, selected_sheet)

        st.download_button(
            label="📥 Скачать готовый файл",
            data=result,
            file_name="Таблица СКС.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
